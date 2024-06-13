import json
import os
import pandas as pd
import openpyxl
from openpyxl.chart import PieChart, LineChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows  
import plotly.express as px

def nacti_json_soubor(soubor):
    """Načte JSON soubor a vrátí jeho obsah."""
    with open(soubor, 'r', encoding='utf-8') as soubor:
        return json.load(soubor)

def preved_json_na_dataframe(data):
    """Převede data z JSON na DataFrame."""
    zpracovana_data = {
        'ID': [],
        'Název': [],
        'Částka': [],
        'Typ financí': [],
        'Název projektu': [],
        'Datum zahájení': [],
        'Datum ukončení': []
    }

    for stranka in data:
        zpracovana_data['ID'].append(stranka['id'])
        zpracovana_data['Název'].append(stranka['name'])
        zpracovana_data['Částka'].append(stranka['amount'])
        zpracovana_data['Typ financí'].append(stranka['financeTypeName'])
        zpracovana_data['Název projektu'].append(stranka['projectName'])
        zpracovana_data['Datum zahájení'].append(stranka['projectStartDate'])
        zpracovana_data['Datum ukončení'].append(stranka['projectEndDate'])

    return pd.DataFrame(zpracovana_data)

def uloz_do_excelu(dataframe, excel_soubor):
    """Uloží DataFrame do Excel souboru včetně grafů."""
    wb = openpyxl.Workbook()
    ws = wb.active

    for radek in dataframe_to_rows(dataframe, index=False, header=True):
        ws.append(radek)

    # Vytvoř koláčový graf
    kolac = PieChart()
    popisky = Reference(ws, min_col=4, min_row=2, max_row=len(dataframe)+1)
    data = Reference(ws, min_col=3, min_row=2, max_row=len(dataframe)+1)
    kolac.add_data(data, titles_from_data=True)
    kolac.set_categories(popisky)
    kolac.title = "Analýza financí - Koláčový graf"
    ws.add_chart(kolac, "G1")

    # lajny pro datum
    line_ws = wb.create_sheet(title="Line Chart Data")

    # Prepare data for line chart including nested project data
    line_data = [
        (
            item['Název'], 
            item['Název projektu'], 
            item['Datum zahájení'], 
            item['Datum ukončení']
        ) for index, item in dataframe.iterrows() 
        if item['Název projektu'] and item['Datum zahájení'] and item['Datum ukončení']
    ]

    line_df = pd.DataFrame(line_data, columns=['FinanceName', 'ProjectName', 'Startdate', 'Enddate'])
    line_df['Startdate'] = pd.to_datetime(line_df['Startdate'])
    line_df['Enddate'] = pd.to_datetime(line_df['Enddate'])

    for r_idx, row in enumerate(dataframe_to_rows(line_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            line_ws.cell(row=r_idx, column=c_idx, value=value)

    # Vytvoř linkový graf do excelu
    line_chart = LineChart()
    line_chart.title = "Timeline projektu"
    line_chart.y_axis.title = "Datum"
    line_chart.x_axis.title = "Zacatek/konec"

    for index, row in line_df.iterrows():
        series = Series(values=Reference(line_ws, min_col=3, min_row=index+2, max_col=4, max_row=index+2),
                        title=row['ProjectName'])
        line_chart.series.append(series)

    line_ws.add_chart(line_chart, "E2")

    # Prepare data for Sunburst chart
    sunburst_ws = wb.create_sheet(title="Sunburst Chart Data")

    sunburst_data = [
        (
            item['Typ financí'],
            item['Název projektu'],
            item['Částka']
        ) for index, item in dataframe.iterrows()
    ]

    sunburst_df = pd.DataFrame(sunburst_data, columns=['FinanceType', 'ProjectName', 'Amount'])

    for r_idx, row in enumerate(dataframe_to_rows(sunburst_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            sunburst_ws.cell(row=r_idx, column=c_idx, value=value)

    # Create Sunburst chart and add it to the new sheet
    sunburst_graf = px.sunburst(sunburst_df, path=['FinanceType', 'ProjectName'], values='Amount', title='Analýza financí - Sunburst graf')
    sunburst_graf.update_traces(textinfo='label+percent entry')
    sunburst_graf.write_image("sunburst_graf.png")
    obrazek = openpyxl.drawing.image.Image("sunburst_graf.png")
    sunburst_ws.add_image(obrazek, 'A1')

    wb.save(excel_soubor)

def main():
    json_soubor = 'finance_analysis.json'
    excel_soubor = 'finance_data.xlsx'

    # Načti JSON soubor
    data = nacti_json_soubor(json_soubor)

    # Převeď JSON data na DataFrame
    df = preved_json_na_dataframe(data)

    # Ulož DataFrame do Excelu s grafy
    uloz_do_excelu(df, excel_soubor)

if __name__ == "__main__":
    main()
